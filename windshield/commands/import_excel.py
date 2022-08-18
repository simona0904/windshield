from pathlib import Path
from windshield.database.windshield import Database
from openpyxl import load_workbook
from windshield.utils import parse_str_to_years, parse_str_to_equipment
from windshield.database.windshield import WindshieldCreateData, EurocodeExists


# functie pt comanda import
def import_excel(file_path: Path, database: Database ):
    wb = load_workbook(file_path)
    sheet1 = wb["Sheet1"]
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        print(row)
        eurocode_without_00 = row[5][0:-2]
        start_year, end_year = parse_str_to_years(row[4])
        sensor, camera, heat = parse_str_to_equipment(eurocode_without_00)
        windshield_create_data = WindshieldCreateData(row[1], row[2], start_year, end_year, sensor, camera, heat, eurocode_without_00)
        try:   
            database.create_windshield(windshield_create_data)
        except EurocodeExists:
            print("Eurocode already exists.")   
        else:
            print("Windshield successfully created.") 
        

